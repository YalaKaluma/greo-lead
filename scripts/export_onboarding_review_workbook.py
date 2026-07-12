from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import desc

from app.db import SessionLocal
from app.models import User


OLD_FLOW_STEPS = [
    ("INITIAL", "Welcome / trigger", "Hey Alfred", "Welcome! I'm Alfred, your AI chief of staff."),
    ("NAME", "Name", "name", "Great to meet you. What do you do?"),
    ("PROFESSION", "Role / profession", "profession", "What's one meaningful goal you want to achieve in the next 6-12 months?"),
    ("GOAL", "First goal", "first_goal", "Why is this important to you?"),
    ("GOAL_WHY", "Goal motivation", "goal_why", "Share a few key tasks - just list them naturally."),
    ("TASKS", "Initial tasks", "tasks_raw", "Which one should you tackle first today?"),
    ("QUICK_WIN", "First quick win", "quick_win", "I've set up your Leadership OS."),
]


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return str(value)


def _safe_step(value: Any) -> str:
    return getattr(value, "value", None) or str(value or "")


def _autosize(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[letter].width = max(12, min(max_len + 2, 60))


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _append_header(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    _style_header(sheet)


def _history_rows(user: User, data: dict[str, Any]) -> list[dict[str, Any]]:
    history = data.get("history") or []
    rows = []
    for index, message in enumerate(history, start=1):
        if not isinstance(message, dict):
            continue
        rows.append(
            {
                "user_id": user.id,
                "user_number": user.phone_number,
                "name": user.name or "",
                "flow_version": data.get("flow_version", ""),
                "prompt_version": data.get("prompt_version", ""),
                "step_number": index,
                "source_step": "history",
                "speaker": message.get("role", ""),
                "prompt_or_question": message.get("content", "") if message.get("role") == "assistant" else "",
                "user_answer": message.get("content", "") if message.get("role") == "user" else "",
                "alfred_response": message.get("content", "") if message.get("role") == "assistant" else "",
                "raw_content": message.get("content", ""),
            }
        )
    return rows


def _legacy_rows(user: User, data: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for index, (step_key, label, data_key, alfred_response) in enumerate(OLD_FLOW_STEPS, start=1):
        answer = data.get(data_key, "")
        if data_key == "name":
            answer = answer or user.name or ""
        elif data_key == "profession":
            answer = answer or user.profession or ""
        if not answer and step_key not in {"INITIAL"}:
            continue
        rows.append(
            {
                "user_id": user.id,
                "user_number": user.phone_number,
                "name": user.name or "",
                "flow_version": data.get("flow_version", "legacy"),
                "prompt_version": data.get("prompt_version", "legacy_whatsapp"),
                "step_number": index,
                "source_step": label,
                "speaker": "user",
                "prompt_or_question": label,
                "user_answer": answer,
                "alfred_response": alfred_response,
                "raw_content": answer,
            }
        )
    return rows


def _conversation_rows(user: User) -> list[dict[str, Any]]:
    data = dict(user.onboarding_data or {})
    rows = _history_rows(user, data)
    if rows:
        return rows
    return _legacy_rows(user, data)


def _is_latest_in_app_onboarding(user: User) -> bool:
    data = dict(user.onboarding_data or {})
    history = data.get("history")
    return (
        data.get("flow_version") == 3
        or data.get("prompt_version") == "onboarding_coach_v3"
        or (isinstance(history, list) and any(isinstance(item, dict) for item in history))
    )


def build_workbook(users: list[User], output: Path) -> int:
    wb = Workbook()
    transcript = wb.active
    transcript.title = "Onboarding Transcript"
    transcript_headers = [
        "User ID",
        "User Number",
        "Name",
        "Flow Version",
        "Prompt Version",
        "Step #",
        "Source Step",
        "Speaker",
        "Prompt / Question",
        "User Answer",
        "Alfred Response",
        "Raw Content",
        "Your Rating (1-5)",
        "Your Feedback",
        "Desired Change",
        "Priority",
        "Reviewed?",
    ]
    _append_header(transcript, transcript_headers)

    step_review = wb.create_sheet("Step Review")
    step_headers = [
        "User ID",
        "User Number",
        "Name",
        "Flow Version",
        "Prompt Version",
        "Step #",
        "Prompt / Alfred Question",
        "User Answer",
        "Alfred Next Response",
        "Extracted Facts At End",
        "Rating: Question Quality",
        "Rating: Response Quality",
        "Feedback",
        "Suggested Better Question / Response",
        "Priority",
        "Reviewed?",
    ]
    _append_header(step_review, step_headers)

    summary = wb.create_sheet("Users Summary")
    summary_headers = [
        "User ID",
        "User Number",
        "Email",
        "Name",
        "Profession",
        "Created At",
        "Onboarding Step",
        "Completed?",
        "Status",
        "Flow Version",
        "Prompt Version",
        "Message Count",
        "Onboarding Data Keys",
        "Generated Payload",
        "Review Notes",
    ]
    _append_header(summary, summary_headers)

    guide = wb.create_sheet("How To Review")
    guide.append(["Field", "How to use it"])
    guide.append(["Your Rating (1-5)", "Use 5 for excellent, 1 for poor. Leave blank until reviewed."])
    guide.append(["Feedback", "Capture what felt wrong, missing, too generic, too long, or especially good."])
    guide.append(["Desired Change", "Write the model behavior you want next time."])
    guide.append(["Priority", "High, Medium, or Low."])
    guide.append(["Reviewed?", "Mark Yes when reviewed."])
    _style_header(guide)

    row_count = 0
    for user in users:
        data = dict(user.onboarding_data or {})
        conversation_rows = _conversation_rows(user)
        for row in conversation_rows:
            transcript.append(
                [
                    row["user_id"],
                    row["user_number"],
                    row["name"],
                    row["flow_version"],
                    row["prompt_version"],
                    row["step_number"],
                    row["source_step"],
                    row["speaker"],
                    row["prompt_or_question"],
                    row["user_answer"],
                    row["alfred_response"],
                    row["raw_content"],
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            row_count += 1

        messages = conversation_rows
        assistant_messages = [message for message in messages if message["speaker"] == "assistant"]
        if assistant_messages:
            for index, message in enumerate(messages):
                if message["speaker"] != "assistant":
                    continue
                next_user = next((m for m in messages[index + 1 :] if m["speaker"] == "user"), {})
                next_assistant = next((m for m in messages[index + 1 :] if m["speaker"] == "assistant"), {})
                step_review.append(
                    [
                        message["user_id"],
                        message["user_number"],
                        message["name"],
                        message["flow_version"],
                        message["prompt_version"],
                        message["step_number"],
                        message["raw_content"],
                        next_user.get("raw_content", ""),
                        next_assistant.get("raw_content", ""),
                        _json_text(data.get("facts")),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )
        else:
            for message in messages:
                step_review.append(
                    [
                        message["user_id"],
                        message["user_number"],
                        message["name"],
                        message["flow_version"],
                        message["prompt_version"],
                        message["step_number"],
                        message["prompt_or_question"],
                        message["user_answer"],
                        message["alfred_response"],
                        _json_text(data.get("facts")),
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ]
                )

        summary.append(
            [
                user.id,
                user.phone_number,
                user.email or "",
                user.name or "",
                user.profession or "",
                user.created_at.isoformat() if user.created_at else "",
                _safe_step(user.onboarding_step),
                bool(user.onboarding_completed),
                data.get("status", ""),
                data.get("flow_version", ""),
                data.get("prompt_version", ""),
                data.get("message_count", ""),
                ", ".join(sorted(data.keys())),
                _json_text(data.get("generated_payload")),
                "",
            ]
        )

    for sheet in wb.worksheets:
        _autosize(sheet)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return row_count


def main() -> None:
    parser = argparse.ArgumentParser(description="Export onboarding conversations to an Excel review workbook.")
    parser.add_argument("--output", default=f"exports/onboarding_review_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--user-number", default=None)
    parser.add_argument("--include-legacy", action="store_true", help="Include older WhatsApp onboarding records.")
    parser.add_argument("--include-empty", action="store_true")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        query = db.query(User)
        if args.user_number:
            query = query.filter(User.phone_number == args.user_number)
        if not args.include_empty:
            query = query.filter(User.onboarding_data.isnot(None))
        users = query.order_by(desc(User.created_at), desc(User.id)).limit(args.limit).all()
        if not args.include_legacy:
            users = [user for user in users if _is_latest_in_app_onboarding(user)]
        output = Path(args.output)
        row_count = build_workbook(users, output)
        print(f"Wrote {output.resolve()} with {len(users)} users and {row_count} transcript rows.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
